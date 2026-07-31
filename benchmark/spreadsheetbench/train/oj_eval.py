#!/usr/bin/env python3
"""OJ-style comparison for SpreadsheetBench outputs.

Cell comparison semantics are ported from upstream evaluation/evaluation.py so
scores stay comparable with the published benchmark. Two deliberate additions:

- When ``answer_position`` has no ``Sheet!`` prefix but the dataset record carries
  an ``answer_sheet`` field, the named sheet is compared (upstream falls back to
  the workbook's first sheet, which can silently compare an untouched sheet).
- Optional formula recalculation: with ``SSB_ALLOW_RECALC=1``, output workbooks
  whose compared range contains uncached formulas are recalculated through
  LibreOffice headless before comparison, so formula-writing solutions are not
  scored as blank cells. OFF by default so scoring matches SkillOpt's evaluator
  (and upstream), which reads cached cell values only and never recalculates.
"""

from __future__ import annotations

import datetime
import os
import shutil
import subprocess
import tempfile
import threading
from pathlib import Path
from typing import Any

import openpyxl

_SOFFICE_CANDIDATES = (
    "soffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
)

# LibreOffice user profiles cannot be shared between concurrent instances; each
# recalc call gets a private profile, and this semaphore caps parallel instances.
_RECALC_SEMAPHORE = threading.Semaphore(int(os.getenv("SSB_RECALC_CONCURRENCY", "2")))


def _find_soffice() -> str | None:
    for candidate in _SOFFICE_CANDIDATES:
        path = shutil.which(candidate) or (candidate if os.path.isfile(candidate) else None)
        if path:
            return path
    return None


def _datetime_to_float(dt: datetime.datetime) -> float:
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def _transform_value(v: Any) -> Any:
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, datetime.time):
        return str(v)[:-3]
    if isinstance(v, datetime.datetime):
        return round(_datetime_to_float(v), 0)
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    return v


def _compare_cell_value(v1: Any, v2: Any) -> bool:
    v1 = _transform_value(v1)
    v2 = _transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) is not type(v2):
        return False
    return v1 == v2


def _col_num2name(n: int) -> str:
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _col_name2num(name: str) -> int:
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def _parse_cell_range(range_str: str) -> tuple[tuple[int, int], tuple[int, int]]:
    start_cell, end_cell = range_str.split(":")

    def split_cell(cell: str) -> tuple[str, str]:
        col, row = "", ""
        for char in cell:
            if char.isdigit():
                row += char
            else:
                col += char
        return col, row

    start_col, start_row = split_cell(start_cell)
    end_col, end_row = split_cell(end_cell)
    return (_col_name2num(start_col), int(start_row)), (_col_name2num(end_col), int(end_row))


def _generate_cell_names(range_str: str) -> list[str]:
    if ":" not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = _parse_cell_range(range_str)
    columns = [_col_num2name(i) for i in range(start_col, end_col + 1)]
    return [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]


def _resolve_ranges(answer_position: str, answer_sheet: Any, default_sheet: str) -> list[tuple[str, str]]:
    """Return (sheet_name, cell_range) pairs for every comma-separated range."""
    resolved: list[tuple[str, str]] = []
    # Some forum-sourced annotations use full-width punctuation (e.g. "G12：J15").
    normalized = str(answer_position).replace("：", ":").replace("，", ",")
    for part in normalized.split(","):
        part = part.strip()
        if not part:
            continue
        if "!" in part:
            sheet_name, cell_range = part.split("!", 1)
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = str(answer_sheet).strip("'") if answer_sheet else default_sheet
            cell_range = part
        resolved.append((sheet_name, cell_range.strip("'")))
    return resolved


def _range_has_uncached_formula(xlsx_path: str | Path, ranges: list[tuple[str, str]]) -> bool:
    """True when a compared cell holds a formula whose cached value is missing."""
    try:
        wb_formula = openpyxl.load_workbook(filename=str(xlsx_path), data_only=False)
        wb_cached = openpyxl.load_workbook(filename=str(xlsx_path), data_only=True)
    except Exception:
        return False
    try:
        for sheet_name, cell_range in ranges:
            if sheet_name not in wb_formula or sheet_name not in wb_cached:
                continue
            ws_formula = wb_formula[sheet_name]
            ws_cached = wb_cached[sheet_name]
            for cell_name in _generate_cell_names(cell_range):
                raw = ws_formula[cell_name].value
                if isinstance(raw, str) and raw.startswith("="):
                    if ws_cached[cell_name].value is None:
                        return True
    finally:
        wb_formula.close()
        wb_cached.close()
    return False


def recalculate_with_libreoffice(xlsx_path: str | Path, timeout: float = 120.0) -> Path | None:
    """Round-trip a workbook through LibreOffice so formula values get cached.

    Returns the path of the recalculated copy (placed next to the original in a
    ``recalc`` subdirectory) or None when recalculation is unavailable/fails.
    """
    soffice = _find_soffice()
    if not soffice:
        return None
    xlsx_path = Path(xlsx_path).resolve()
    out_dir = xlsx_path.parent / "recalc"
    out_dir.mkdir(exist_ok=True)
    with _RECALC_SEMAPHORE:
        with tempfile.TemporaryDirectory(prefix="ssb_soffice_profile_") as profile_dir:
            try:
                subprocess.run(
                    [
                        soffice,
                        "--headless",
                        "--norestore",
                        f"-env:UserInstallation=file://{profile_dir}",
                        "--convert-to",
                        "xlsx:Calc MS Excel 2007 XML",
                        "--outdir",
                        str(out_dir),
                        str(xlsx_path),
                    ],
                    capture_output=True,
                    timeout=timeout,
                    check=False,
                )
            except Exception:
                return None
    recalced = out_dir / xlsx_path.name
    return recalced if recalced.exists() else None


def recalc_enabled() -> bool:
    """Whether the judge recalculates uncached formulas before comparing.

    Defaults to off so scoring matches SkillOpt's evaluator (cached values
    only). The system prompt in the rollout executor keys off this too, so the
    agent is never told formulas are acceptable when the judge would read them
    as blank.
    """
    return os.getenv("SSB_ALLOW_RECALC") == "1"


def compare_workbooks(
    gt_file: str | Path,
    proc_file: str | Path,
    answer_position: str,
    answer_sheet: Any = None,
    *,
    allow_recalc: bool | None = None,
) -> tuple[bool, str]:
    """Compare one produced workbook against the ground-truth answer workbook."""
    if allow_recalc is None:
        allow_recalc = recalc_enabled()
    gt_file = str(gt_file)
    proc_file = str(proc_file)
    if not os.path.exists(proc_file):
        return False, "output file not produced"
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
    except Exception as exc:
        return False, f"failed to open answer workbook: {exc}"
    ranges = _resolve_ranges(answer_position, answer_sheet, wb_gt.sheetnames[0])

    if allow_recalc and _range_has_uncached_formula(proc_file, ranges):
        recalced = recalculate_with_libreoffice(proc_file)
        if recalced is not None:
            proc_file = str(recalced)

    try:
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as exc:
        return False, f"failed to open output workbook: {exc}"

    try:
        for sheet_name, cell_range in ranges:
            if sheet_name not in wb_gt.sheetnames:
                # Fall back to upstream behaviour for malformed annotations.
                sheet_name = wb_gt.sheetnames[0]
            if sheet_name not in wb_proc.sheetnames:
                return False, f"worksheet not found in output: {sheet_name}"
            ws_gt = wb_gt[sheet_name]
            ws_proc = wb_proc[sheet_name]
            for cell_name in _generate_cell_names(cell_range):
                gt_value = ws_gt[cell_name].value
                proc_value = ws_proc[cell_name].value
                if not _compare_cell_value(gt_value, proc_value):
                    return False, (
                        f"value mismatch at {sheet_name}!{cell_name}: "
                        f"expected {gt_value!r}, got {proc_value!r}"
                    )
    finally:
        wb_gt.close()
        wb_proc.close()
    return True, ""
